from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime, timedelta
from apps.core_app.models import DailyWorkLog, WorkLogAttendance, Worker
from django.db.models import Sum


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_excel(request):
    date_from_str = request.query_params.get('date_from')
    date_to_str = request.query_params.get('date_to')
    filter_type = request.query_params.get('filter', 'all')
    filter_id = request.query_params.get('filter_id')
    
    date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else datetime.now().date() - timedelta(days=30)
    date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else datetime.now().date()
    
    logs = DailyWorkLog.objects.filter(log_date__range=[date_from, date_to]).select_related('block', 'crop', 'work_type', 'season')
    
    if filter_type == 'block' and filter_id:
        logs = logs.filter(block_id=filter_id)
    elif filter_type == 'crop' and filter_id:
        logs = logs.filter(crop_id=filter_id)
    elif filter_type == 'worker' and filter_id:
        work_log_ids = WorkLogAttendance.objects.filter(worker_id=filter_id).values_list('work_log_id', flat=True)
        logs = logs.filter(id__in=work_log_ids)
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Work Logs'
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['Date', 'Block', 'Crop', 'Season', 'Row', 'Work Type', 'Work Detail', 'Male', 'Female', 'Total', 'Hours', 'Status', 'Workers', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    for row_idx, log in enumerate(logs, 2):
        worker_names = ', '.join([
            a.worker.name for a in 
            WorkLogAttendance.objects.filter(work_log=log).select_related('worker')
        ])
        
        ws.cell(row=row_idx, column=1, value=log.log_date)
        ws.cell(row=row_idx, column=2, value=log.block.name)
        ws.cell(row=row_idx, column=3, value=log.crop.name)
        ws.cell(row=row_idx, column=4, value=log.season.name if log.season else '')
        ws.cell(row=row_idx, column=5, value=log.row_number or '')
        ws.cell(row=row_idx, column=6, value=log.work_type.name)
        ws.cell(row=row_idx, column=7, value=log.work_detail)
        ws.cell(row=row_idx, column=8, value=log.male_labour_count)
        ws.cell(row=row_idx, column=9, value=log.female_labour_count)
        ws.cell(row=row_idx, column=10, value=log.total_labour_count)
        ws.cell(row=row_idx, column=11, value=float(log.hours_worked or 0))
        ws.cell(row=row_idx, column=12, value=log.status)
        ws.cell(row=row_idx, column=13, value=worker_names)
        ws.cell(row=row_idx, column=14, value=log.notes)
        
        for col in range(1, 15):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    col_widths = {'A': 12, 'B': 12, 'C': 15, 'D': 14, 'E': 6, 'F': 18, 'G': 20, 'H': 8, 'I': 8, 'J': 8, 'K': 8, 'L': 10, 'M': 30, 'N': 30}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    summary_ws = wb.create_sheet('Summary')
    summary_ws['A1'] = f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    summary_ws['A2'] = f"Period: {date_from} to {date_to}"
    summary_ws['A3'] = f"Total Work Items: {logs.count()}"
    
    total_labour = sum(log.total_labour_count for log in logs)
    total_male = sum(log.male_labour_count for log in logs)
    total_female = sum(log.female_labour_count for log in logs)
    summary_ws['A4'] = f"Total Labour Days: {total_labour}"
    summary_ws['A5'] = f"Total Male: {total_male}"
    summary_ws['A6'] = f"Total Female: {total_female}"
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return FileResponse(
        output,
        as_attachment=True,
        filename=f"work_logs_{date_from}_{date_to}.xlsx",
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
